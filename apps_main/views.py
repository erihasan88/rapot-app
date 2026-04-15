from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib import messages
from .models import (
    User, Student, Teacher, ClassLevel, Subject, Grade, AcademicYear,
    Attendance, P5Project, TeacherNote, Attitude, SchoolProfile, TeacherSubject, SubjectActivity
)
from django.db.models import Count, Avg, Sum, Max, Min
from decimal import Decimal, InvalidOperation
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime, date

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

def _xlsx_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return response

def _parse_excel_date(value):
    if value is None or value == '':
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None

def auto_description_from_numeric_grade(value):
    if value is None or value == '':
        return ''
    try:
        score = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return ''
    if score >= 90:
        return "Siswa telah menguasai materi dengan sangat baik, mampu menerapkan konsep secara mandiri, serta menunjukkan pemahaman yang sangat kuat."
    if score >= 80:
        return "Siswa telah mencapai tujuan pembelajaran dengan baik, memahami konsep, namun masih perlu sedikit peningkatan dalam penerapannya."
    if score >= 70:
        return "Siswa memiliki pemahaman yang cukup, namun masih memerlukan bimbingan dalam menerapkan konsep."
    return "Siswa belum mencapai tujuan pembelajaran dan memerlukan bimbingan intensif untuk meningkatkan pemahaman."

def compute_report_grade(daily_test, midterm, final_exam, activity_score=None):
    def to_decimal(v):
        if v is None or v == '':
            return None
        try:
            return Decimal(str(v))
        except (InvalidOperation, ValueError):
            return None

    d = to_decimal(daily_test)
    m = to_decimal(midterm)
    f = to_decimal(final_exam)
    a = to_decimal(activity_score)
    if d is None and m is None and f is None and a is None:
        return None

    weights = []
    values = []
    if d is not None:
        weights.append(Decimal('0.40'))
        values.append(d)
    if m is not None:
        weights.append(Decimal('0.30'))
        values.append(m)
    if f is not None:
        weights.append(Decimal('0.30'))
        values.append(f)

    total_weight = sum(weights)
    base = None
    if total_weight != 0:
        scaled = [w / total_weight for w in weights]
        base = sum(v * w for v, w in zip(values, scaled))

    result = None
    if base is not None and a is not None:
        result = (base * Decimal('0.90')) + (a * Decimal('0.10'))
    elif base is not None:
        result = base
    else:
        result = a

    return result.quantize(Decimal('1'))

def is_admin(user):
    return user.is_authenticated and user.role == 'ADMIN'

def is_teacher(user):
    return user.is_authenticated and user.role == 'TEACHER'

def is_student(user):
    return user.is_authenticated and user.role == 'STUDENT'

@login_required
def dashboard(request):
    user = request.user
    context = {
        'page_title': 'Dashboard',
    }
    
    if user.role == 'ADMIN':
        context.update({
            'student_count': Student.objects.count(),
            'teacher_count': Teacher.objects.count(),
            'class_count': ClassLevel.objects.count(),
            'subject_count': Subject.objects.count(),
            'recent_students': Student.objects.all().order_by('-id')[:5],
            'paket_b_count': Student.objects.filter(class_level__program='PAKET_B').count(),
            'paket_c_count': Student.objects.filter(class_level__program='PAKET_C').count(),
            'active_year': AcademicYear.objects.filter(is_active=True).first(),
        })
        return render(request, 'apps_main/admin_dashboard.html', context)
    
    elif user.role == 'TEACHER':
        if request.method == 'POST':
            if 'photo' in request.FILES:
                user.photo = request.FILES['photo']
                user.save()
                messages.success(request, "Foto profil berhasil diperbarui!")
            return redirect('dashboard')

        teacher = get_object_or_404(Teacher, user=user)
        context.update({
            'teacher': teacher,
            'subjects': teacher.teachersubject_set.all(),
        })
        return render(request, 'apps_main/teacher_dashboard.html', context)
    
    elif user.role == 'STUDENT':
        student = get_object_or_404(Student, user=user)
        active_year = AcademicYear.objects.filter(is_active=True).first()
        grades = Grade.objects.filter(student=student, academic_year=active_year).select_related('subject')
        activities = SubjectActivity.objects.filter(student=student, academic_year=active_year).select_related('subject')

        rows_by_subject = {}
        for g in grades:
            rows_by_subject[g.subject_id] = {
                'subject': g.subject,
                'daily_test_grade': g.daily_test_grade,
                'activity_score': None,
                'activity_note': '',
            }
        for a in activities:
            row = rows_by_subject.get(a.subject_id)
            if not row:
                row = {
                    'subject': a.subject,
                    'daily_test_grade': None,
                    'activity_score': None,
                    'activity_note': '',
                }
                rows_by_subject[a.subject_id] = row
            row['activity_score'] = a.activity_score
            row['activity_note'] = a.activity_note or ''

        subject_rows = sorted(rows_by_subject.values(), key=lambda r: (r['subject'].name or '').lower())
        context.update({
            'student': student,
            'active_year': active_year,
            'subject_rows': subject_rows,
        })
        return render(request, 'apps_main/student_dashboard.html', context)
    
    return redirect('login')

@user_passes_test(is_admin)
def manage_students(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        full_name = request.POST.get('full_name')
        email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')
        nis = request.POST.get('nis')
        nisn = request.POST.get('nisn')
        class_id = request.POST.get('class_id')
        gender = request.POST.get('gender')
        birth_place = request.POST.get('birth_place')
        birth_date = request.POST.get('birth_date')
        parent_name = request.POST.get('parent_name')
        address = request.POST.get('address')
        
        try:
            if student_id:
                # Update existing student
                student = get_object_or_404(Student, id=student_id)
                user = student.user
                
                # Check if new username is already taken by ANOTHER user
                if User.objects.exclude(id=user.id).filter(username=username).exists():
                    messages.error(request, f"Gagal: Username '{username}' sudah digunakan oleh orang lain.")
                    return redirect('manage_students')
                
                # Check if new email is already taken by ANOTHER user
                if email and User.objects.exclude(id=user.id).filter(email=email).exists():
                    messages.error(request, f"Gagal: Email '{email}' sudah terdaftar di sistem.")
                    return redirect('manage_students')

                user.username = username
                user.email = email
                user.first_name = full_name
                user.last_name = ''
                if password:
                    user.set_password(password)
                
                if 'photo' in request.FILES:
                    user.photo = request.FILES['photo']
                    
                user.save()
                
                class_level = ClassLevel.objects.get(id=class_id) if class_id else None
                student.nis = nis
                student.nisn = nisn
                student.class_level = class_level
                student.gender = gender
                student.birth_place = birth_place
                student.birth_date = birth_date if birth_date else None
                student.parent_name = parent_name
                student.address = address
                student.save()
                messages.success(request, f"Data siswa {full_name} berhasil diperbarui!")
            else:
                # Check if username or email already exists for new user
                if User.objects.filter(username=username).exists():
                    messages.error(request, f"Gagal: Username '{username}' sudah digunakan. Silakan gunakan username lain.")
                    return redirect('manage_students')
                
                if email and User.objects.filter(email=email).exists():
                    messages.error(request, f"Gagal: Email '{email}' sudah terdaftar. Gunakan email lain.")
                    return redirect('manage_students')

                # Create User first
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=full_name,
                    role='STUDENT'
                )
                user.last_name = ''
                user.save()
                
                if 'photo' in request.FILES:
                    user.photo = request.FILES['photo']
                    user.save()
                
                # Create Student profile
                class_level = ClassLevel.objects.get(id=class_id) if class_id else None
                Student.objects.create(
                    user=user,
                    nis=nis,
                    nisn=nisn,
                    class_level=class_level,
                    gender=gender,
                    birth_place=birth_place,
                    birth_date=birth_date if birth_date else None,
                    parent_name=parent_name,
                    address=address
                )
                messages.success(request, f"Siswa {full_name} berhasil ditambahkan!")
        except Exception as e:
            messages.error(request, f"Gagal memproses data siswa: {str(e)}")
            
        return redirect('manage_students')

    students = Student.objects.all().select_related('class_level', 'user').order_by('class_level__name', 'user__first_name')
    classes = ClassLevel.objects.all().order_by('name')
    return render(request, 'apps_main/manage_students.html', {
        'students': students, 
        'classes': classes,
        'page_title': 'Manajemen Siswa'
    })

@user_passes_test(is_admin)
def export_students_excel(request):
    if Workbook is None:
        messages.error(request, "Fitur Excel belum tersedia. Install dulu: pip install openpyxl")
        return redirect('manage_students')

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = [
        "nisn", "nis", "full_name", "username", "email",
        "class_name", "program", "gender", "birth_place", "birth_date",
        "parent_name", "address",
    ]
    ws.append(headers)

    students = Student.objects.all().select_related('user', 'class_level').order_by('class_level__name', 'user__first_name')
    for s in students:
        ws.append([
            s.nisn,
            s.nis or "",
            s.user.get_full_name(),
            s.user.username,
            s.user.email or "",
            s.class_level.name if s.class_level else "",
            s.class_level.get_program_display() if s.class_level else "",
            s.gender,
            s.birth_place or "",
            s.birth_date.strftime("%Y-%m-%d") if s.birth_date else "",
            s.parent_name or "",
            s.address or "",
        ])

    return _xlsx_response(wb, "students.xlsx")


@user_passes_test(is_admin)
def import_students_excel(request):
    if request.method != 'POST':
        return redirect('manage_students')
    if load_workbook is None:
        messages.error(request, "Fitur Excel belum tersedia. Install dulu: pip install openpyxl")
        return redirect('manage_students')

    file = request.FILES.get('excel_file')
    if not file:
        messages.error(request, "Pilih file Excel terlebih dahulu.")
        return redirect('manage_students')

    wb = load_workbook(filename=file, data_only=True)
    ws = wb["Students"] if "Students" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, "File Excel kosong.")
        return redirect('manage_students')

    header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
    idx = {name: header.index(name) for name in header if name}

    def get(row, key, default=""):
        i = idx.get(key)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return default if v is None else v

    created_count = 0
    updated_count = 0
    error_lines = []

    for row_num, row in enumerate(rows[1:], start=2):
        nisn = str(get(row, "nisn", "")).strip()
        full_name = str(get(row, "full_name", "")).strip()
        username = str(get(row, "username", "")).strip()

        if not nisn or not full_name or not username:
            continue

        email = str(get(row, "email", "")).strip()
        nis = str(get(row, "nis", "")).strip()
        class_name = str(get(row, "class_name", "")).strip()
        gender = str(get(row, "gender", "L")).strip() or "L"
        birth_place = str(get(row, "birth_place", "")).strip()
        birth_date = _parse_excel_date(get(row, "birth_date", ""))
        parent_name = str(get(row, "parent_name", "")).strip()
        address = str(get(row, "address", "")).strip()
        password = str(get(row, "password", "")).strip() if "password" in idx else ""

        class_level = None
        if class_name:
            class_level = ClassLevel.objects.filter(name__iexact=class_name).first()
            if not class_level:
                error_lines.append(f"Baris {row_num}: Kelas '{class_name}' tidak ditemukan.")
                continue

        try:
            student = Student.objects.filter(nisn=nisn).select_related('user').first()
            if student:
                user = student.user
                if User.objects.exclude(id=user.id).filter(username=username).exists():
                    error_lines.append(f"Baris {row_num}: Username '{username}' sudah dipakai.")
                    continue
                if email and User.objects.exclude(id=user.id).filter(email=email).exists():
                    error_lines.append(f"Baris {row_num}: Email '{email}' sudah terdaftar.")
                    continue
                user.username = username
                user.email = email
                user.first_name = full_name
                if password:
                    user.set_password(password)
                user.save()

                student.nis = nis
                student.class_level = class_level
                student.gender = gender if gender in ["L", "P"] else student.gender
                student.birth_place = birth_place
                student.birth_date = birth_date
                student.parent_name = parent_name
                student.address = address
                student.save()
                updated_count += 1
            else:
                if User.objects.filter(username=username).exists():
                    error_lines.append(f"Baris {row_num}: Username '{username}' sudah dipakai.")
                    continue
                if email and User.objects.filter(email=email).exists():
                    error_lines.append(f"Baris {row_num}: Email '{email}' sudah terdaftar.")
                    continue
                if Student.objects.filter(nisn=nisn).exists():
                    error_lines.append(f"Baris {row_num}: NISN '{nisn}' sudah ada.")
                    continue

                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password or None,
                    first_name=full_name,
                    role='STUDENT'
                )
                Student.objects.create(
                    user=user,
                    nis=nis,
                    nisn=nisn,
                    class_level=class_level,
                    gender=gender if gender in ["L", "P"] else "L",
                    birth_place=birth_place,
                    birth_date=birth_date,
                    parent_name=parent_name,
                    address=address,
                )
                created_count += 1
        except Exception as e:
            error_lines.append(f"Baris {row_num}: {str(e)}")

    if created_count or updated_count:
        messages.success(request, f"Import siswa selesai. Baru: {created_count}, Update: {updated_count}.")
    if error_lines:
        messages.error(request, "Sebagian baris gagal diproses: " + " | ".join(error_lines[:10]))

    return redirect('manage_students')

@user_passes_test(is_admin)
def manage_teachers(request):
    if request.method == 'POST':
        teacher_id = request.POST.get('teacher_id')
        full_name = request.POST.get('full_name')
        email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')
        nip = request.POST.get('nip')
        homeroom_class_id = request.POST.get('homeroom_class_id')
        
        # Get assignments (Subject-Class pairs)
        subject_ids = request.POST.getlist('subjects[]')
        class_ids = request.POST.getlist('classes[]')
        
        try:
            if teacher_id:
                # Update existing teacher
                teacher = get_object_or_404(Teacher, id=teacher_id)
                user = teacher.user
                
                # Check unique username/email
                if User.objects.exclude(id=user.id).filter(username=username).exists():
                    messages.error(request, f"Gagal: Username '{username}' sudah digunakan.")
                    return redirect('manage_teachers')
                
                if email and User.objects.exclude(id=user.id).filter(email=email).exists():
                    messages.error(request, f"Gagal: Email '{email}' sudah terdaftar.")
                    return redirect('manage_teachers')

                user.username = username
                user.email = email
                user.first_name = full_name
                if password:
                    user.set_password(password)
                user.save()
                
                teacher.nip = nip
                teacher.save()

                # Update assignments
                TeacherSubject.objects.filter(teacher=teacher).delete()
                for s_id, c_id in zip(subject_ids, class_ids):
                    if s_id and c_id:
                        TeacherSubject.objects.create(
                            teacher=teacher,
                            subject_id=s_id,
                            class_level_id=c_id
                        )

                # Update wali kelas (homeroom)
                if homeroom_class_id:
                    ClassLevel.objects.filter(homeroom_teacher=teacher).exclude(id=homeroom_class_id).update(homeroom_teacher=None)
                    class_obj = get_object_or_404(ClassLevel, id=homeroom_class_id)
                    class_obj.homeroom_teacher = teacher
                    class_obj.save()
                else:
                    ClassLevel.objects.filter(homeroom_teacher=teacher).update(homeroom_teacher=None)

                messages.success(request, f"Data guru {full_name} berhasil diperbarui!")
            else:
                # Check unique username/email
                if User.objects.filter(username=username).exists():
                    messages.error(request, f"Gagal: Username '{username}' sudah digunakan.")
                    return redirect('manage_teachers')
                
                if email and User.objects.filter(email=email).exists():
                    messages.error(request, f"Gagal: Email '{email}' sudah terdaftar.")
                    return redirect('manage_teachers')

                # Create User
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=full_name,
                    role='TEACHER'
                )
                
                # Create Teacher profile
                teacher = Teacher.objects.create(user=user, nip=nip)

                # Create assignments
                for s_id, c_id in zip(subject_ids, class_ids):
                    if s_id and c_id:
                        TeacherSubject.objects.create(
                            teacher=teacher,
                            subject_id=s_id,
                            class_level_id=c_id
                        )

                # Set wali kelas (homeroom) if provided
                if homeroom_class_id:
                    class_obj = get_object_or_404(ClassLevel, id=homeroom_class_id)
                    class_obj.homeroom_teacher = teacher
                    class_obj.save()

                messages.success(request, f"Guru {full_name} berhasil ditambahkan!")
        except Exception as e:
            messages.error(request, f"Gagal memproses data guru: {str(e)}")
            
        return redirect('manage_teachers')

    teachers = Teacher.objects.all().order_by('user__first_name')
    subjects = Subject.objects.all().order_by('name')
    classes = ClassLevel.objects.all().order_by('name')
    return render(request, 'apps_main/manage_teachers.html', {
        'teachers': teachers, 
        'subjects': subjects,
        'classes': classes,
        'page_title': 'Data Guru'
    })

@user_passes_test(is_admin)
def export_teachers_excel(request):
    if Workbook is None:
        messages.error(request, "Fitur Excel belum tersedia. Install dulu: pip install openpyxl")
        return redirect('manage_teachers')

    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"

    ws.append(["username", "full_name", "email", "nip", "homeroom_class"])
    teachers = Teacher.objects.all().select_related('user').order_by('user__first_name')
    for t in teachers:
        homeroom = t.homeroom_classes.first()
        ws.append([
            t.user.username,
            t.user.get_full_name(),
            t.user.email or "",
            t.nip or "",
            homeroom.name if homeroom else "",
        ])

    ws2 = wb.create_sheet("TeacherAssignments")
    ws2.append(["teacher_username", "subject_code", "subject_name", "class_name"])
    assignments = TeacherSubject.objects.all().select_related('teacher__user', 'subject', 'class_level').order_by('teacher__user__first_name')
    for ts in assignments:
        ws2.append([
            ts.teacher.user.username,
            ts.subject.code,
            ts.subject.name,
            ts.class_level.name,
        ])

    return _xlsx_response(wb, "teachers.xlsx")


@user_passes_test(is_admin)
def import_teachers_excel(request):
    if request.method != 'POST':
        return redirect('manage_teachers')
    if load_workbook is None:
        messages.error(request, "Fitur Excel belum tersedia. Install dulu: pip install openpyxl")
        return redirect('manage_teachers')

    file = request.FILES.get('excel_file')
    if not file:
        messages.error(request, "Pilih file Excel terlebih dahulu.")
        return redirect('manage_teachers')

    wb = load_workbook(filename=file, data_only=True)
    ws = wb["Teachers"] if "Teachers" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, "File Excel kosong.")
        return redirect('manage_teachers')

    header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
    idx = {name: header.index(name) for name in header if name}

    def get(row, key, default=""):
        i = idx.get(key)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return default if v is None else v

    created_count = 0
    updated_count = 0
    error_lines = []

    for row_num, row in enumerate(rows[1:], start=2):
        username = str(get(row, "username", "")).strip()
        full_name = str(get(row, "full_name", "")).strip()
        nip = str(get(row, "nip", "")).strip()
        if not username or not full_name:
            continue

        email = str(get(row, "email", "")).strip()
        homeroom_class = str(get(row, "homeroom_class", "")).strip()
        password = str(get(row, "password", "")).strip() if "password" in idx else ""

        try:
            teacher = None
            if nip:
                teacher = Teacher.objects.filter(nip=nip).select_related('user').first()
            if not teacher:
                teacher = Teacher.objects.filter(user__username=username).select_related('user').first()

            if teacher:
                user = teacher.user
                if User.objects.exclude(id=user.id).filter(username=username).exists():
                    error_lines.append(f"Baris {row_num}: Username '{username}' sudah dipakai.")
                    continue
                if email and User.objects.exclude(id=user.id).filter(email=email).exists():
                    error_lines.append(f"Baris {row_num}: Email '{email}' sudah terdaftar.")
                    continue

                user.username = username
                user.email = email
                user.first_name = full_name
                if password:
                    user.set_password(password)
                user.save()

                if nip:
                    teacher.nip = nip
                teacher.save()
                updated_count += 1
            else:
                if User.objects.filter(username=username).exists():
                    error_lines.append(f"Baris {row_num}: Username '{username}' sudah dipakai.")
                    continue
                if email and User.objects.filter(email=email).exists():
                    error_lines.append(f"Baris {row_num}: Email '{email}' sudah terdaftar.")
                    continue
                if nip and Teacher.objects.filter(nip=nip).exists():
                    error_lines.append(f"Baris {row_num}: NIP '{nip}' sudah terdaftar.")
                    continue

                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password or None,
                    first_name=full_name,
                    role='TEACHER'
                )
                teacher = Teacher.objects.create(user=user, nip=nip or None)
                created_count += 1

            if homeroom_class:
                class_obj = ClassLevel.objects.filter(name__iexact=homeroom_class).first()
                if not class_obj:
                    error_lines.append(f"Baris {row_num}: Kelas wali '{homeroom_class}' tidak ditemukan.")
                else:
                    class_obj.homeroom_teacher = teacher
                    class_obj.save()
        except Exception as e:
            error_lines.append(f"Baris {row_num}: {str(e)}")

    if "TeacherAssignments" in wb.sheetnames:
        ws_a = wb["TeacherAssignments"]
        rows_a = list(ws_a.iter_rows(values_only=True))
        if rows_a:
            header_a = [str(x).strip().lower() if x is not None else "" for x in rows_a[0]]
            idx_a = {name: header_a.index(name) for name in header_a if name}

            def get_a(row, key, default=""):
                i = idx_a.get(key)
                if i is None or i >= len(row):
                    return default
                v = row[i]
                return default if v is None else v

            for row_num, row in enumerate(rows_a[1:], start=2):
                t_username = str(get_a(row, "teacher_username", "")).strip()
                subject_code = str(get_a(row, "subject_code", "")).strip()
                class_name = str(get_a(row, "class_name", "")).strip()
                if not t_username or not subject_code or not class_name:
                    continue
                teacher = Teacher.objects.filter(user__username=t_username).first()
                subject = Subject.objects.filter(code__iexact=subject_code).first()
                class_level = ClassLevel.objects.filter(name__iexact=class_name).first()
                if not teacher or not subject or not class_level:
                    continue
                TeacherSubject.objects.get_or_create(teacher=teacher, subject=subject, class_level=class_level)

    if created_count or updated_count:
        messages.success(request, f"Import guru selesai. Baru: {created_count}, Update: {updated_count}.")
    if error_lines:
        messages.error(request, "Sebagian baris gagal diproses: " + " | ".join(error_lines[:10]))

    return redirect('manage_teachers')

@user_passes_test(is_admin)
def manage_classes(request):
    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        name = request.POST.get('name')
        program = request.POST.get('program')
        phase = request.POST.get('phase')
        
        try:
            if class_id:
                # Update existing class
                class_obj = get_object_or_404(ClassLevel, id=class_id)
                class_obj.name = name
                class_obj.program = program
                class_obj.phase = phase
                class_obj.save()
                messages.success(request, f"Kelas {name} berhasil diperbarui!")
            else:
                # Create new class
                ClassLevel.objects.create(name=name, program=program, phase=phase)
                messages.success(request, f"Kelas {name} berhasil ditambahkan!")
        except Exception as e:
            messages.error(request, f"Gagal memproses data kelas: {str(e)}")
            
        return redirect('manage_classes')

    classes = ClassLevel.objects.all().order_by('program', 'name')
    return render(request, 'apps_main/manage_classes.html', {'classes': classes, 'page_title': 'Manajemen Kelas'})

@user_passes_test(is_admin)
def manage_subjects(request):
    if request.method == 'POST':
        subject_id = request.POST.get('subject_id')
        name = request.POST.get('name')
        
        try:
            if subject_id:
                subject = get_object_or_404(Subject, id=subject_id)
                subject.name = name
                subject.save()
                messages.success(request, f"Mata pelajaran {name} berhasil diperbarui!")
            else:
                Subject.objects.create(name=name)
                messages.success(request, f"Mata pelajaran {name} berhasil ditambahkan!")
        except Exception as e:
            messages.error(request, f"Gagal memproses data: {str(e)}")
        return redirect('manage_subjects')

    subjects = Subject.objects.all().order_by('name')
    return render(request, 'apps_main/manage_subjects.html', {'subjects': subjects, 'page_title': 'Mata Pelajaran'})

@user_passes_test(is_admin)
def manage_academic_years(request):
    if request.method == 'POST':
        year_id = request.POST.get('year_id')
        year = request.POST.get('year')
        semester = request.POST.get('semester')
        is_active = request.POST.get('is_active') == 'on'
        
        try:
            if is_active:
                AcademicYear.objects.all().update(is_active=False)
                
            if year_id:
                obj = get_object_or_404(AcademicYear, id=year_id)
                obj.year = year
                obj.semester = semester
                obj.is_active = is_active
                obj.save()
                messages.success(request, f"Tahun ajaran {year} diperbarui!")
            else:
                AcademicYear.objects.create(year=year, semester=semester, is_active=is_active)
                messages.success(request, f"Tahun ajaran {year} ditambahkan!")
        except Exception as e:
            messages.error(request, f"Gagal memproses data: {str(e)}")
        return redirect('manage_academic_years')

    years = AcademicYear.objects.all().order_by('-year', '-semester')
    return render(request, 'apps_main/manage_academic_years.html', {'years': years, 'page_title': 'Tahun Ajaran'})

@user_passes_test(is_admin)
def manage_grades(request):
    if request.method == 'POST':
        grade_id = request.POST.get('grade_id')
        numeric_grade = request.POST.get('numeric_grade')
        description = request.POST.get('description')
        teacher_id = request.POST.get('teacher_id')
        next_url = request.POST.get('next') or request.get_full_path()

        try:
            grade = get_object_or_404(Grade, id=grade_id)
            if numeric_grade is not None and numeric_grade != '':
                grade.numeric_grade = numeric_grade
            if description and description.strip():
                grade.description = description.strip()
            else:
                grade.description = auto_description_from_numeric_grade(numeric_grade or grade.numeric_grade)
            if teacher_id:
                grade.teacher_id = teacher_id
            grade.save()
            messages.success(request, "Nilai berhasil diperbarui!")
        except Exception as e:
            messages.error(request, f"Gagal memperbarui nilai: {str(e)}")

        return redirect(next_url)

    selected_year = request.GET.get('year')
    selected_subject = request.GET.get('subject')
    
    grades = Grade.objects.all().select_related('student__user', 'subject', 'teacher__user', 'academic_year')
    
    if selected_year:
        grades = grades.filter(academic_year_id=selected_year)
    if selected_subject:
        grades = grades.filter(subject_id=selected_subject)
        
    context = {
        'grades': grades.order_by('-academic_year__year', 'student__user__first_name'),
        'academic_years': AcademicYear.objects.all().order_by('-year', '-semester'),
        'subjects': Subject.objects.all().order_by('name'),
        'teachers': Teacher.objects.all().select_related('user').order_by('user__first_name'),
        'selected_year': selected_year,
        'selected_subject': selected_subject,
        'page_title': 'Data Nilai'
    }
    return render(request, 'apps_main/manage_grades.html', context)

@user_passes_test(is_admin)
def manage_settings(request):
    school = SchoolProfile.objects.first()
    if not school:
        school = SchoolProfile.objects.create(name="PKBM Darul Ulum Agrabinta")

    if request.method == 'POST':
        admin_email = (request.POST.get('admin_email') or '').strip()
        current_password = request.POST.get('current_password') or ''
        new_password = request.POST.get('new_password') or ''
        confirm_password = request.POST.get('confirm_password') or ''

        school.name = request.POST.get('name')
        school.address = request.POST.get('address')
        school.phone = request.POST.get('phone')
        school.email = request.POST.get('email')
        school.principal_name = request.POST.get('principal_name')
        school.nip = request.POST.get('nip')
        school.theme_color = request.POST.get('theme_color')
        
        if 'logo' in request.FILES:
            school.logo = request.FILES['logo']
        if 'signature' in request.FILES:
            school.signature = request.FILES['signature']
            
        try:
            if admin_email:
                if User.objects.exclude(id=request.user.id).filter(username=admin_email).exists():
                    messages.error(request, f"Gagal: Email '{admin_email}' sudah digunakan sebagai username.")
                    return redirect('manage_settings')
                if User.objects.exclude(id=request.user.id).filter(email=admin_email).exists():
                    messages.error(request, f"Gagal: Email '{admin_email}' sudah terdaftar.")
                    return redirect('manage_settings')
                request.user.username = admin_email
                request.user.email = admin_email
                request.user.save()

            if new_password or confirm_password or current_password:
                if not new_password or not confirm_password or not current_password:
                    messages.error(request, "Gagal: Lengkapi password saat ini, password baru, dan ulangi password.")
                    return redirect('manage_settings')
                if new_password != confirm_password:
                    messages.error(request, "Gagal: Password baru dan konfirmasi tidak sama.")
                    return redirect('manage_settings')
                if not request.user.check_password(current_password):
                    messages.error(request, "Gagal: Password saat ini salah.")
                    return redirect('manage_settings')
                request.user.set_password(new_password)
                request.user.save()
                update_session_auth_hash(request, request.user)

            school.save()
            messages.success(request, "Pengaturan berhasil diperbarui!")
        except Exception as e:
            messages.error(request, f"Gagal memproses pengaturan: {str(e)}")
            return redirect('manage_settings')

        return redirect('manage_settings')

    color_options = [
        '#1e295b', '#0f172a', '#312e81', '#4338ca', 
        '#1e40af', '#0369a1', '#0f766e', '#166534'
    ]

    return render(request, 'apps_main/manage_settings.html', {
        'school': school,
        'page_title': 'Pengaturan Sekolah',
        'color_options': color_options
    })

@user_passes_test(is_teacher)
def input_grades(request):
    teacher = get_object_or_404(Teacher, user=request.user)
    active_year = AcademicYear.objects.filter(is_active=True).first()
    ts_id = request.GET.get('ts_id')
    
    students = []
    if ts_id:
        ts = get_object_or_404(TeacherSubject, id=ts_id, teacher=teacher)
        students = Student.objects.filter(class_level=ts.class_level)
        for s in students:
            s.current_grade = Grade.objects.filter(student=s, subject=ts.subject, academic_year=active_year).first()
            s.current_activity = SubjectActivity.objects.filter(student=s, subject=ts.subject, academic_year=active_year).first()
            
    if request.method == 'POST':
        ts = get_object_or_404(TeacherSubject, id=ts_id, teacher=teacher)
        for s in students:
            daily = request.POST.get(f'daily_{s.id}')
            midterm = request.POST.get(f'midterm_{s.id}')
            final_exam = request.POST.get(f'final_{s.id}')
            desc = request.POST.get(f'desc_{s.id}')
            activity_score = request.POST.get(f'activity_{s.id}')
            activity_note = request.POST.get(f'activity_note_{s.id}')

            report_grade = compute_report_grade(daily, midterm, final_exam, activity_score)
            if report_grade is not None:
                final_desc = (desc or '').strip() or auto_description_from_numeric_grade(report_grade)

                Grade.objects.update_or_create(
                    student=s, subject=ts.subject, academic_year=active_year,
                    defaults={
                        'daily_test_grade': daily or None,
                        'midterm_grade': midterm or None,
                        'final_exam_grade': final_exam or None,
                        'numeric_grade': report_grade,
                        'description': final_desc,
                        'teacher': teacher
                    }
                )

            if (activity_score and str(activity_score).strip() != '') or (activity_note and str(activity_note).strip() != ''):
                SubjectActivity.objects.update_or_create(
                    student=s, subject=ts.subject, academic_year=active_year,
                    defaults={
                        'activity_score': activity_score or None,
                        'activity_note': (activity_note or '').strip(),
                        'teacher': teacher
                    }
                )
        messages.success(request, "Nilai berhasil disimpan!")
        return redirect(f"{request.path}?ts_id={ts_id}")

    class_stats = None
    if ts_id and students:
        ts = get_object_or_404(TeacherSubject, id=ts_id, teacher=teacher)
        qs = Grade.objects.filter(
            academic_year=active_year,
            subject=ts.subject,
            student__class_level=ts.class_level
        )
        agg = qs.aggregate(
            total=Sum('numeric_grade'),
            avg=Avg('numeric_grade'),
            max=Max('numeric_grade'),
            min=Min('numeric_grade'),
            count=Count('id')
        )
        avg_val = agg['avg'] or Decimal('0')
        class_stats = {
            'count': agg['count'] or 0,
            'total': agg['total'] or Decimal('0'),
            'avg': avg_val,
            'absorption': avg_val,
            'max': agg['max'] or Decimal('0'),
            'min': agg['min'] or Decimal('0'),
        }

    return render(request, 'apps_main/input_grades.html', {
        'teacher': teacher,
        'students': students,
        'page_title': 'Input Nilai',
        'class_stats': class_stats,
    })

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

@user_passes_test(is_teacher)
def student_detail_input(request, student_id):
    teacher = get_object_or_404(Teacher, user=request.user)
    student = get_object_or_404(Student, id=student_id)
    active_year = AcademicYear.objects.filter(is_active=True).first()
    
    attendance, _ = Attendance.objects.get_or_create(student=student, academic_year=active_year)
    attitude, _ = Attitude.objects.get_or_create(student=student, academic_year=active_year)
    note, _ = TeacherNote.objects.get_or_create(student=student, academic_year=active_year)
    
    if request.method == 'POST':
        # Attendance
        attendance.sakit = request.POST.get('sakit', 0)
        attendance.izin = request.POST.get('izin', 0)
        attendance.alpha = request.POST.get('alpha', 0)
        attendance.save()
        
        # Attitude
        attitude.score = request.POST.get('attitude_score')
        attitude.description = request.POST.get('attitude_desc')
        attitude.save()
        
        # Note
        note.notes = request.POST.get('notes')
        note.save()
        
        # P5 (Single project for demo simplicity)
        p5_title = request.POST.get('p5_title')
        if p5_title:
            P5Project.objects.update_or_create(
                student=student, academic_year=active_year, title=p5_title,
                defaults={
                    'description': request.POST.get('p5_desc'),
                    'score': request.POST.get('p5_score')
                }
            )
            
        messages.success(request, f"Data {student.user.get_full_name()} berhasil diperbarui!")
        return redirect('input_grades')

    context = {
        'student': student,
        'attendance': attendance,
        'attitude': attitude,
        'note': note,
        'p5': P5Project.objects.filter(student=student, academic_year=active_year).first(),
        'page_title': f'Detail Rapor: {student.user.get_full_name()}'
    }
    return render(request, 'apps_main/student_detail_input.html', context)

@user_passes_test(is_student)
def view_report(request):
    student = get_object_or_404(Student, user=request.user)
    active_year = AcademicYear.objects.filter(is_active=True).first()
    grades = Grade.objects.filter(student=student, academic_year=active_year)
    attendance = Attendance.objects.filter(student=student, academic_year=active_year).first()
    p5 = P5Project.objects.filter(student=student, academic_year=active_year)
    note = TeacherNote.objects.filter(student=student, academic_year=active_year).first()
    attitude = Attitude.objects.filter(student=student, academic_year=active_year).first()
    school = SchoolProfile.objects.first()
    
    if request.GET.get('format') == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapor_{student.user.username}_{active_year.year.replace("/", "_")}.pdf"'
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Header
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, fontSize=14, spaceAfter=10)
        elements.append(Paragraph(school.name.upper(), title_style))
        elements.append(Paragraph(school.address, styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(f"LAPORAN HASIL BELAJAR (RAPOR)", title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Student Info
        info_data = [
            [f"Nama Siswa: {student.user.get_full_name().upper()}", f"Kelas: {student.class_level.name}"],
            [f"NISN: {student.nisn}", f"Semester: {active_year.get_semester_display()}"],
            [f"Program: {student.class_level.get_program_display()}", f"Tahun Ajaran: {active_year.year}"]
        ]
        info_table = Table(info_data, colWidths=[10*cm, 7*cm])
        info_table.setStyle(TableStyle([
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Grades Table
        elements.append(Paragraph("A. NILAI AKADEMIK", styles['Heading3']))
        grade_data = [['No', 'Mata Pelajaran', 'Nilai', 'Capaian Kompetensi']]
        for i, g in enumerate(grades, 1):
            grade_data.append([i, g.subject.name, int(g.numeric_grade), Paragraph(g.description, styles['Normal'])])
        
        grade_table = Table(grade_data, colWidths=[1*cm, 5*cm, 2*cm, 9*cm])
        grade_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('ALIGN', (2,0), (2,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(grade_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Attendance
        elements.append(Paragraph("B. KEHADIRAN", styles['Heading3']))
        att_data = [
            ['Sakit', f"{attendance.sakit if attendance else 0} hari"],
            ['Izin', f"{attendance.izin if attendance else 0} hari"],
            ['Tanpa Keterangan', f"{attendance.alpha if attendance else 0} hari"]
        ]
        att_table = Table(att_data, colWidths=[4*cm, 3*cm])
        att_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTSIZE', (0,0), (-1,-1), 9),
        ]))
        elements.append(att_table)
        elements.append(Spacer(1, 0.5*cm))

        # Attitude
        if attitude:
            elements.append(Paragraph("C. SIKAP & KARAKTER", styles['Heading3']))
            elements.append(Paragraph(f"<b>Predikat:</b> {attitude.score}", styles['Normal']))
            elements.append(Paragraph(f"<b>Deskripsi:</b> {attitude.description}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))

        # P5
        if p5:
            elements.append(Paragraph("D. PROJEK P5", styles['Heading3']))
            for p in p5:
                elements.append(Paragraph(f"<b>{p.title}</b>", styles['Normal']))
                elements.append(Paragraph(f"Predikat: {p.score}", styles['Normal']))
                elements.append(Paragraph(p.description, styles['Normal']))
                elements.append(Spacer(1, 0.2*cm))
            elements.append(Spacer(1, 0.3*cm))

        # Notes
        if note:
            elements.append(Paragraph("E. CATATAN GURU", styles['Heading3']))
            elements.append(Paragraph(f"<i>\"{note.notes}\"</i>", styles['Normal']))
            elements.append(Spacer(1, 1*cm))

        # Signature
        principal_images = []
        if school.signature and hasattr(school.signature, 'path'):
            try:
                principal_images.append(Image(school.signature.path, width=5*cm, height=2*cm))
            except Exception:
                principal_images.append(Spacer(1, 2*cm))
        else:
            principal_images.append(Spacer(1, 2*cm))

        sig_data = [
            ["Orang Tua/Wali", "", f"Agrabinta, {date.today().strftime('%d %B %Y')}"],
            ["", "", "Kepala PKBM Darul Ulum"],
            ["", "", principal_images],
            ["(..........................)", "", f"<b>{school.principal_name}</b>"],
            ["", "", f"NIP. {school.nip if school.nip else '-'}"],
        ]
        sig_table = Table(sig_data, colWidths=[6*cm, 5*cm, 6*cm])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
        ]))
        elements.append(sig_table)
        
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    context = {
        'student': student,
        'year': active_year,
        'grades': grades,
        'attendance': attendance,
        'p5': p5,
        'note': note,
        'attitude': attitude,
        'school': school,
        'page_title': 'Rapor Siswa',
        'today': date.today()
    }
    return render(request, 'apps_main/view_report.html', context)
